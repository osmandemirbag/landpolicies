"""
Create an Excel file of Weimar Republic (1919-1933) parliamentary bills
related to agricultural protectionism, rural credits, mortgage credits,
and related beneficial agricultural legislation.

Sources:
- Reichsgesetzblatt (RGBl) archives
- JSTOR: Agrarian Protectionism in the Weimar Republic (https://www.jstor.org/stable/260171)
- Oxford Handbook of the Weimar Republic, Agriculture and Rural Society chapter
- Foreign Affairs: The Crisis in German Agriculture (1932)
- Wikipedia: Reich Settlement Law, Reich Ministry of Food and Agriculture
- Gesetze-im-Internet: RSiedlG, PachtkredG
- Historisches Lexikon Bayerns: Osthilfe, 1926-1937
- 1000dokumente.de: Verordnung über die Errichtung der Deutschen Rentenbank
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Bill data ────────────────────────────────────────────────────────────────

bills = [
    # (year, german_name, english_name, date_enacted, rgbl_citation, category,
    #  status, description, key_provisions, beneficiaries, initiating_body)

    # ─── 1919 ───
    (1919,
     "Reichssiedlungsgesetz (RSiedlG)",
     "Reich Settlement Law",
     "11 August 1919",
     "RGBl. 1919 S. 1429",
     "Land Reform / Settlement",
     "Passed",
     "Facilitated the creation of new agricultural settlements by enabling "
     "non-profit settlement companies to acquire and redistribute large estate land. "
     "Aimed to break up large East Elbian estates and create small/medium farms.",
     "Right of pre-emption for settlement companies; limited compulsory acquisition "
     "of estates over 100 ha; government-backed land purchases; subsidised loans for settlers.",
     "Smallholders, war veterans, landless agricultural workers",
     "Weimar National Assembly (Nationalversammlung)"),

    (1919,
     "Verordnung über Landarbeiter (Landarbeiterordnung)",
     "Agricultural Workers' Ordinance",
     "24 January 1919",
     "RGBl. 1919 S. 111",
     "Agricultural Labour",
     "Passed",
     "Established legal protections and regulated employment conditions for "
     "agricultural labourers in the aftermath of WWI, granting them collective "
     "bargaining rights and minimum contract standards.",
     "8-hour working day for farm workers; collective bargaining rights; "
     "minimum employment contract standards; workplace safety provisions.",
     "Agricultural workers, rural labourers",
     "Council of the People's Deputies (Rat der Volksbeauftragten)"),

    (1919,
     "Gesetz über die Sozialisierung (Sozialisierungsgesetz)",
     "Socialisation Law (Agricultural Provisions)",
     "23 March 1919",
     "RGBl. 1919 S. 341",
     "Land Reform / Socialisation",
     "Passed",
     "General framework law enabling the socialisation of suitable economic "
     "enterprises, including agricultural estates. Provided constitutional basis "
     "for potential nationalisation of land but was never broadly applied to agriculture.",
     "Framework for socialisation of enterprises including agricultural estates; "
     "Article 155 of Weimar Constitution mandated land use for the common good.",
     "General public, landless rural population",
     "Weimar National Assembly (Nationalversammlung)"),

    (1919,
     "SPD-Entwurf eines Bodenreformgesetzes",
     "SPD Land Reform Bill (1919)",
     "1919",
     "N/A – Reichstag bill",
     "Land Reform",
     "Rejected / Not passed",
     "Radical proposal by the Social Democratic Party to expropriate large Junker "
     "estates (especially in East Elbia) with compensation and redistribute the "
     "land to small farmers and landless workers. Blocked by conservative and "
     "centrist parties in the National Assembly.",
     "Compulsory expropriation of estates exceeding a threshold; redistribution "
     "to smallholders; state-managed land funds.",
     "Landless workers, small farmers",
     "SPD (Sozialdemokratische Partei Deutschlands)"),

    # ─── 1920 ───
    (1920,
     "Reichsheimstättengesetz",
     "Reich Homestead Law",
     "10 May 1920",
     "RGBl. 1920 I S. 962",
     "Land Reform / Settlement",
     "Passed",
     "Promoted the establishment of small homesteads (Heimstätten) for families, "
     "especially returning soldiers, workers, and low-income groups. Homesteads "
     "were protected against forced sale and excessive mortgage burden.",
     "Creation of protected homesteads; restrictions on forced sale and "
     "mortgage foreclosure; special credit provisions for homestead owners.",
     "Low-income families, war veterans, rural settlers",
     "Reichstag"),

    (1920,
     "KPD-Antrag auf entschädigungslose Enteignung des Großgrundbesitzes",
     "KPD Motion for Uncompensated Expropriation of Large Estates (1920)",
     "1920",
     "N/A – Reichstag motion",
     "Land Reform",
     "Rejected",
     "Communist Party motion calling for the uncompensated expropriation of all "
     "large estates and their transfer to public or cooperative ownership. "
     "Overwhelmingly rejected by the Reichstag due to opposition from centre, "
     "right, and liberal parties.",
     "Complete expropriation without compensation of all large estates; "
     "transfer to state or cooperatives.",
     "Landless workers, agricultural cooperatives",
     "KPD (Kommunistische Partei Deutschlands)"),

    # ─── 1921 ───
    (1921,
     "Reichssiedlungsgesetz – Erste Novelle",
     "Reich Settlement Law – First Amendment",
     "7 June 1923",
     "RGBl. 1923 I S. 364",
     "Land Reform / Settlement",
     "Passed",
     "First major amendment to the 1919 Reich Settlement Law, tightening the "
     "obligations of large landowners to provide land for settlement and adjusting "
     "financial provisions for the inflationary period.",
     "Strengthened pre-emption rights; adjusted compensation mechanisms "
     "for hyperinflation; extended settlement company powers.",
     "Smallholders, new settlers",
     "Reichstag"),

    # ─── 1922 ───
    (1922,
     "Gesetz über die Aufhebung der Fideikommisse",
     "Law on the Abolition of Entailed Estates (Fideicommissa)",
     "6 July 1938 (final dissolution); process began 1919–1922",
     "Various state laws; Weimar Constitution Art. 155",
     "Land Reform",
     "Passed (gradual)",
     "The Weimar Constitution (Art. 155) mandated the dissolution of entailed "
     "estates (Fideikommisse) – large hereditary estates that could not be divided "
     "or sold. Prussian and other state laws during 1920-1922 began implementation, "
     "though full dissolution took decades.",
     "Abolition of primogeniture entail on agricultural estates; conversion to "
     "freely transferable property; facilitation of land market.",
     "Land market participants, potential smallholders",
     "Weimar Constitution / State (Länder) legislatures"),

    # ─── 1923 ───
    (1923,
     "Verordnung über die Errichtung der Deutschen Rentenbank",
     "Ordinance on the Establishment of the German Rentenbank",
     "15 October 1923",
     "RGBl. 1923 I S. 963",
     "Rural Credit / Monetary Stabilisation",
     "Passed",
     "Created the Deutsche Rentenbank as a mortgage bank to back the new "
     "Rentenmark currency during the hyperinflation crisis. The bank was capitalised "
     "by mortgages on agricultural and industrial property, providing critical "
     "monetary stabilisation and laying the foundation for agricultural credit.",
     "Establishment of the Rentenbank backed by land mortgages; issuance of "
     "Rentenmark; stabilisation of the currency and rural finance.",
     "All German farmers (as mortgage-providers and currency beneficiaries)",
     "Reich Government (Reichsregierung) – emergency decree"),

    (1923,
     "Dritte Verordnung zur Durchführung der Rentenmarkverordnung",
     "Third Implementing Ordinance for the Rentenmark Regulation",
     "20 November 1923",
     "RGBl. 1923 I S. 1133",
     "Rural Credit / Monetary Stabilisation",
     "Passed",
     "Detailed implementing regulations for the Rentenbank and Rentenmark, "
     "specifying how agricultural mortgages would be registered and how "
     "Rentenmark would circulate in the rural economy.",
     "Technical provisions for agricultural mortgage registration; "
     "Rentenmark circulation rules; Rentenbank operational guidelines.",
     "Farmers, agricultural creditors",
     "Reich Government (Reichsregierung)"),

    # ─── 1924 ───
    (1924,
     "Verordnung über die Goldmarkbilanz der Landwirtschaft",
     "Ordinance on Gold-Mark Balance Sheets for Agriculture",
     "1924",
     "RGBl. 1924 I",
     "Agricultural Finance / Debt",
     "Passed",
     "Required agricultural enterprises to restate their balance sheets in "
     "gold marks after the hyperinflation, establishing a new accounting basis "
     "for farm assets, debts, and mortgages.",
     "Mandatory gold-mark restatement of agricultural balance sheets; "
     "new valuation framework for farm assets and liabilities.",
     "All agricultural enterprises",
     "Reich Government (Reichsregierung)"),

    # ─── 1925 ───
    (1925,
     "Gesetz über die Errichtung der Deutschen Rentenbank-Kreditanstalt",
     "Law on the Establishment of the German Rentenbank Credit Institution",
     "18 July 1925",
     "RGBl. 1925 I S. 145",
     "Rural Credit",
     "Passed",
     "Created the Rentenbank-Kreditanstalt as a specialised institution for "
     "providing medium- and long-term agricultural credit at low interest rates, "
     "enabling farmers to refinance debts, invest in improvements, and access "
     "capital unavailable from private banks.",
     "Medium/long-term agricultural loans at reduced rates; refinancing of "
     "farm debts; support for agricultural investment and modernisation.",
     "All farmers, especially small and medium-sized operations",
     "Reichstag"),

    (1925,
     "Aufwertungsgesetz",
     "Revaluation Law",
     "16 July 1925",
     "RGBl. 1925 I S. 117",
     "Mortgage / Debt Regulation",
     "Passed",
     "Partially restored the value of mortgages and debts that had been wiped "
     "out by hyperinflation. Agricultural mortgages were revalued at 25% of "
     "their pre-inflation gold-mark value, balancing the interests of creditors "
     "and debtor-farmers.",
     "Revaluation of pre-inflation mortgages at 25% of gold-mark value; "
     "special provisions for agricultural and rural mortgages; phased repayment schedules.",
     "Agricultural mortgage holders (both creditors and debtors)",
     "Reichstag"),

    (1925,
     "Zolltarifgesetz (Neufassung)",
     "Customs Tariff Law (New Version)",
     "25 August 1925",
     "RGBl. 1925 I S. 289",
     "Agricultural Protectionism / Tariffs",
     "Passed",
     "Reintroduced protective tariffs on agricultural imports (especially grain, "
     "meat, and dairy) after the post-war liberal trade period. Marked Germany's "
     "return to agricultural protectionism to shield domestic farmers from "
     "falling world prices.",
     "Increased import duties on grain, livestock, meat, dairy products; "
     "tariff schedule for agricultural commodities; administered by customs authorities.",
     "Grain farmers, livestock farmers, dairy producers",
     "Reichstag"),

    # ─── 1926 ───
    (1926,
     "Pachtkreditgesetz (PachtkredG)",
     "Tenant Credit Law",
     "9 July 1926",
     "RGBl. 1926 I S. 399",
     "Rural Credit",
     "Passed",
     "Enabled tenant farmers to use their farming inventory (livestock, "
     "machinery, tools) as collateral for loans, without transferring possession. "
     "Addressed the chronic credit shortage among non-landowning farmers.",
     "Pledge rights on agricultural inventory for tenant farmers; "
     "court-registered security interests; expanded credit access without land ownership.",
     "Tenant farmers (Pächter)",
     "Reichstag"),

    (1926,
     "Futtermittelgesetz",
     "Feed Law",
     "22 December 1926",
     "RGBl. 1926 I S. 525",
     "Agricultural Regulation / Protectionism",
     "Passed",
     "Regulated the production, trade, and quality standards of animal feed. "
     "Included import restrictions on foreign feed products to protect domestic "
     "feed grain producers and ensure livestock feed quality.",
     "Quality standards for animal feed; import controls on foreign feed; "
     "labelling and inspection requirements.",
     "Livestock farmers, domestic feed producers",
     "Reichstag"),

    (1926,
     "Ostpreußenhilfe (Gesetz über Hilfsmaßnahmen für Ostpreußen)",
     "East Prussia Aid Law",
     "1926",
     "Prussian state law / Reich administrative measures",
     "Rural Credit / Regional Aid",
     "Passed",
     "First major aid programme for the economically distressed eastern "
     "provinces, particularly East Prussia. Provided emergency credit, tax "
     "relief, and transport subsidies to agricultural estates suffering from "
     "geographic isolation and market disadvantage.",
     "Emergency credit lines; tax relief for eastern agricultural estates; "
     "transport subsidies; infrastructure improvement funding.",
     "Eastern German farmers and estate owners (especially East Prussia)",
     "Prussian state government / Reichsregierung"),

    (1926,
     "Volksbegehren zur Enteignung der Fürstenvermögen (incl. Agrarland)",
     "People's Initiative for Expropriation of Princely Properties (incl. Agricultural Land)",
     "20 June 1926 (referendum)",
     "N/A – referendum",
     "Land Reform",
     "Rejected (referendum failed)",
     "SPD/KPD-backed popular referendum to expropriate without compensation the "
     "properties of former German ruling houses, including vast agricultural "
     "estates. The referendum attracted 14.5 million yes votes but failed to "
     "reach the required 50% threshold of eligible voters.",
     "Uncompensated expropriation of all former royal/princely properties "
     "including agricultural land; transfer to public ownership.",
     "General public, potential land reform beneficiaries",
     "SPD / KPD (referendum initiative)"),

    # ─── 1927 ───
    (1927,
     "Zollerhöhung für landwirtschaftliche Erzeugnisse",
     "Tariff Increase on Agricultural Products",
     "1927",
     "RGBl. 1927 I (various decrees)",
     "Agricultural Protectionism / Tariffs",
     "Passed",
     "Series of increases in import duties on key agricultural products "
     "including rye, wheat, barley, and livestock. Part of ongoing protectionist "
     "policy to stabilise farm incomes as world agricultural prices declined.",
     "Higher specific duties on grain imports; increased tariffs on meat "
     "and dairy imports; sliding-scale tariff adjustments.",
     "German grain and livestock farmers",
     "Reichstag / Reichsregierung"),

    # ─── 1928 ───
    (1928,
     "Erweiterung der Ostpreußenhilfe (Zweites Osthilfe-Programm)",
     "Extension of East Prussia Aid (Second Eastern Aid Programme)",
     "1928",
     "Administrative decrees / Reich budget appropriations",
     "Rural Credit / Regional Aid",
     "Passed",
     "Expanded the 1926 East Prussia aid programme to include broader eastern "
     "regions (Pomerania, Silesia, Brandenburg border areas). Increased credit "
     "volumes and introduced agricultural debt restructuring measures.",
     "Expanded geographic scope; increased credit volumes; debt restructuring "
     "provisions; agricultural investment subsidies.",
     "Farmers in eastern German provinces",
     "Reichsregierung"),

    (1928,
     "Gesetz über die Rentenbank-Kreditanstalt (Novelle)",
     "Amendment to the Rentenbank Credit Institution Law",
     "31 March 1928",
     "RGBl. 1928 I",
     "Rural Credit",
     "Passed",
     "Amended the 1925 founding law of the Rentenbank-Kreditanstalt to expand "
     "its lending capacity, adjust its capital structure, and broaden the types "
     "of agricultural loans it could offer.",
     "Expanded lending authority; adjusted capital requirements; broadened "
     "eligible loan types for agricultural modernisation.",
     "All agricultural borrowers",
     "Reichstag"),

    # ─── 1929 ───
    (1929,
     "Zolltarifnovelle (Erhöhung der Agrarzölle)",
     "Tariff Amendment (Increase of Agricultural Duties)",
     "1929",
     "RGBl. 1929 I (various decrees)",
     "Agricultural Protectionism / Tariffs",
     "Passed",
     "Further substantial increases in agricultural import tariffs in response "
     "to the onset of the Great Depression and collapsing world commodity prices. "
     "Grain, sugar, meat, and dairy tariffs were raised significantly.",
     "Major increases in grain tariffs (especially rye and wheat); higher "
     "duties on sugar, meat, and dairy; tightened import quota system.",
     "All agricultural producers, especially grain farmers",
     "Reichstag / Reichsregierung"),

    (1929,
     "Agrarkreditgesetz",
     "Agricultural Credit Law",
     "1929",
     "RGBl. 1929 I",
     "Rural Credit",
     "Passed",
     "Eased access to agricultural credit by providing government-backed "
     "guarantees for farm loans and establishing new mechanisms for emergency "
     "credit during the agricultural depression.",
     "Government-guaranteed agricultural loans; emergency credit facilities; "
     "simplified application procedures for farm credit.",
     "All farmers, especially those facing insolvency",
     "Reichstag"),

    # ─── 1930 ───
    (1930,
     "Einfuhrscheinverordnung (Verordnung über Einfuhrscheine)",
     "Import Certificate Regulation",
     "1930",
     "RGBl. 1930 I",
     "Agricultural Protectionism / Trade Control",
     "Passed",
     "Introduced a mandatory import certificate (Einfuhrschein) system for "
     "agricultural products, especially grains. Importers were required to obtain "
     "government-issued certificates, strictly controlling the volume of foreign "
     "agricultural goods entering Germany.",
     "Mandatory import certificates for agricultural products; volume controls "
     "on grain imports; administrative allocation of import rights.",
     "Domestic grain and agricultural producers",
     "Reichsregierung (emergency decree)"),

    (1930,
     "Getreidegesetz (Verordnung zur Regelung des Getreideverkehrs)",
     "Grain Law (Regulation of Grain Trade)",
     "1930",
     "RGBl. 1930 I",
     "Agricultural Protectionism / Market Regulation",
     "Passed",
     "Regulated domestic grain trade, established state purchase obligations, "
     "and introduced minimum price mechanisms to stabilise the grain market "
     "during the agricultural depression.",
     "State grain purchase obligations; minimum price guarantees for rye and "
     "wheat; regulation of grain storage and distribution.",
     "Grain farmers, rural communities",
     "Reichsregierung (emergency decree)"),

    (1930,
     "Notverordnung des Reichspräsidenten zur Sicherung von Wirtschaft und Finanzen "
     "(Agrarbestimmungen)",
     "Presidential Emergency Decree for Securing the Economy and Finance "
     "(Agricultural Provisions)",
     "26 July 1930",
     "RGBl. 1930 I S. 311",
     "Agricultural Finance / Emergency Measures",
     "Passed (emergency decree)",
     "First of the major Brüning-era emergency decrees with specific agricultural "
     "provisions: tax relief for farmers, temporary moratoriums on farm debt "
     "execution, and increased tariff protection.",
     "Tax relief for agriculture; temporary stay on farm foreclosures; "
     "increased agricultural tariffs; emergency credit provisions.",
     "All farmers, especially those facing insolvency",
     "Reichspräsident (Article 48 of Weimar Constitution)"),

    # ─── 1931 ───
    (1931,
     "Osthilfegesetz (Gesetz über die Hilfsmaßnahmen für die notleidenden "
     "Gebiete des Ostens)",
     "Eastern Aid Law (Law on Aid Measures for the Distressed Eastern Territories)",
     "31 March 1931",
     "RGBl. 1931 I S. 109",
     "Rural Credit / Regional Aid",
     "Passed",
     "Major federal law providing comprehensive financial aid to agricultural "
     "estates in eastern Germany. Total programme reached ~2.5 billion RM. "
     "Controversial for disproportionately benefiting large Junker estates over "
     "small farmers.",
     "Large-scale credit and loan guarantees; tax relief; transport subsidies; "
     "debt restructuring for eastern estates; settlement programme funding.",
     "Eastern German estate owners (primarily Junkers), also some small farmers",
     "Reichstag"),

    (1931,
     "Dritte Notverordnung des Reichspräsidenten zur Sicherung von Wirtschaft "
     "und Finanzen (Agrarbestimmungen)",
     "Third Presidential Emergency Decree for Securing the Economy and Finance "
     "(Agricultural Provisions)",
     "6 October 1931",
     "RGBl. 1931 I S. 537",
     "Agricultural Finance / Emergency Measures",
     "Passed (emergency decree)",
     "Comprehensive emergency decree with major agricultural provisions including "
     "agricultural price support mechanisms, further tariff increases, and "
     "emergency credit for distressed farms.",
     "Agricultural price support programmes; further tariff increases; "
     "emergency farm credit; extended foreclosure moratoriums.",
     "All farmers, especially debt-distressed operations",
     "Reichspräsident (Article 48 of Weimar Constitution)"),

    (1931,
     "Verordnung des Reichspräsidenten über Maßnahmen zur Erhaltung der "
     "Leistungsfähigkeit der deutschen Landwirtschaft",
     "Presidential Decree on Measures to Maintain the Productivity of "
     "German Agriculture",
     "1931",
     "RGBl. 1931 I",
     "Agricultural Protectionism / Market Regulation",
     "Passed (emergency decree)",
     "Emergency decree introducing price support mechanisms (Preisstützung) "
     "for key agricultural commodities including rye, potatoes, and sugar beet. "
     "Government purchased surpluses to maintain floor prices.",
     "Minimum price guarantees for rye, potatoes, sugar beet; government "
     "surplus purchases; market intervention mechanisms.",
     "Crop farmers, especially rye and potato producers",
     "Reichspräsident (Article 48 of Weimar Constitution)"),

    (1931,
     "Schuldnernotrecht (Verordnung über die Regelung der landwirtschaftlichen "
     "Schuldverhältnisse)",
     "Emergency Debtors' Law (Regulation of Agricultural Debt Relations)",
     "29 December 1931",
     "RGBl. 1931 I S. 745",
     "Agricultural Debt / Mortgage Relief",
     "Passed (emergency decree)",
     "Imposed a moratorium on agricultural debt repayments and mortgage "
     "foreclosures for farmers facing insolvency. Protected many farmers from "
     "losing their land during the Depression but criticised for merely "
     "postponing the debt crisis.",
     "Moratorium on farm debt repayments; suspension of mortgage foreclosures; "
     "mandatory debt mediation procedures; interest rate caps on agricultural loans.",
     "Indebted farmers facing foreclosure",
     "Reichspräsident (Article 48 of Weimar Constitution)"),

    (1931,
     "Milch- und Fettgesetz (Verordnung zur Regelung des Milch- und Fettmarktes)",
     "Milk and Fat Law (Regulation of the Milk and Fat Market)",
     "1931",
     "RGBl. 1931 I",
     "Agricultural Protectionism / Market Regulation",
     "Passed (emergency decree)",
     "Regulated the dairy and fats market, introduced minimum prices for milk, "
     "and imposed tariffs on imported butter and margarine to protect domestic "
     "dairy farmers.",
     "Minimum milk prices; import tariffs on butter, margarine, and fats; "
     "quality standards for dairy products; market regulation mechanisms.",
     "Dairy farmers, domestic butter producers",
     "Reichsregierung (emergency decree)"),

    # ─── 1932 ───
    (1932,
     "Verordnung des Reichspräsidenten über die Entschuldung der Landwirtschaft "
     "(Landwirtschaftliches Entschuldungsverfahren)",
     "Presidential Decree on Agricultural Debt Relief "
     "(Agricultural Debt Settlement Procedure)",
     "27 March 1932",
     "RGBl. 1932 I S. 141",
     "Agricultural Debt / Mortgage Relief",
     "Passed (emergency decree)",
     "Extended and deepened the December 1931 debt moratorium. Established formal "
     "debt settlement procedures (Entschuldungsverfahren) for agricultural "
     "enterprises, allowing for partial debt write-downs and mandatory "
     "creditor-debtor negotiations.",
     "Formal agricultural debt settlement tribunals; partial debt write-downs; "
     "mandatory creditor negotiations; continued foreclosure moratorium.",
     "Heavily indebted farmers and agricultural enterprises",
     "Reichspräsident (Article 48 of Weimar Constitution)"),

    (1932,
     "Notverordnung des Reichspräsidenten über Maßnahmen auf dem Gebiete der "
     "Viehwirtschaft",
     "Presidential Emergency Decree on Measures in the Livestock Economy",
     "1932",
     "RGBl. 1932 I",
     "Agricultural Protectionism / Livestock",
     "Passed (emergency decree)",
     "Introduced import quotas and increased tariffs on livestock and meat "
     "imports. Established government intervention mechanisms for the livestock "
     "market to stabilise prices for cattle, pigs, and poultry farmers.",
     "Import quotas on livestock and meat; increased customs duties; government "
     "market intervention for livestock prices.",
     "Livestock farmers, domestic meat producers",
     "Reichspräsident (Article 48 of Weimar Constitution)"),

    (1932,
     "Gesetz über Förderung der Kleinsiedlung und des Kleingartenwesens "
     "(Kleinsiedlungsgesetz)",
     "Law on the Promotion of Small Settlements and Allotment Gardens "
     "(Small Settlement Law)",
     "1932",
     "RGBl. 1932 I",
     "Land Reform / Settlement",
     "Passed",
     "Promoted the creation of small agricultural settlements and allotment "
     "gardens for urban unemployed and rural poor. Part of the government's "
     "work-creation programmes during the Depression. Provided state subsidies "
     "for settlement construction.",
     "State subsidies for small settlement construction; land allocation for "
     "allotment gardens; building material assistance; integration with "
     "work-creation schemes.",
     "Unemployed workers, urban poor, rural settlers",
     "Reichstag / Reichsregierung"),

    (1932,
     "Mühlengesetz (Verordnung zur Regelung des Mühlengewerbes)",
     "Mill Law (Regulation of the Milling Industry)",
     "1932",
     "RGBl. 1932 I",
     "Agricultural Market Regulation",
     "Passed (emergency decree)",
     "Regulated the milling industry to protect domestic grain producers by "
     "establishing mandatory domestic grain purchase quotas for millers and "
     "restricting the use of imported grain in flour production.",
     "Mandatory domestic grain purchase quotas for mills; mixing ratios for "
     "domestic/imported grain; price controls on flour; milling licences.",
     "Domestic grain farmers, milling industry",
     "Reichsregierung (emergency decree)"),

    (1932,
     "SPD-Entwurf eines umfassenden Bodenreformgesetzes",
     "SPD Comprehensive Land Reform Bill (1932)",
     "1932",
     "N/A – Reichstag bill",
     "Land Reform",
     "Rejected / Not debated",
     "SPD proposal for comprehensive land reform including progressive land "
     "taxation, mandatory subdivision of estates over 1,000 ha, and expanded "
     "settlement programmes. Never received substantive debate due to the "
     "political paralysis of the late Weimar Reichstag.",
     "Progressive land taxation; mandatory estate subdivision; expanded "
     "settlement funding; cooperative farming support.",
     "Small farmers, landless agricultural workers",
     "SPD (Sozialdemokratische Partei Deutschlands)"),

    # ─── 1933 ───
    (1933,
     "Verordnung des Reichspräsidenten über die landwirtschaftliche "
     "Schuldenregelung (Erweiterung)",
     "Presidential Decree on Agricultural Debt Regulation (Extension)",
     "23 January 1933",
     "RGBl. 1933 I",
     "Agricultural Debt / Mortgage Relief",
     "Passed (emergency decree)",
     "Final extension of the agricultural debt moratorium and debt settlement "
     "procedures before the Nazi seizure of power. Continued protection of "
     "indebted farmers from foreclosure.",
     "Extended moratorium on farm debt repayments; continued foreclosure "
     "protection; expanded debt mediation.",
     "Indebted farmers",
     "Reichspräsident (Article 48 of Weimar Constitution)"),

    (1933,
     "DNVP/Stahlhelm-Entwurf eines Gesetzes zur Erhaltung des Bauerntums",
     "DNVP/Stahlhelm Bill for the Preservation of the Peasantry",
     "January 1933",
     "N/A – Reichstag bill (precursor to Nazi Erbhofgesetz)",
     "Agricultural Protection / Inheritance",
     "Not passed (superseded by Nazi Reichserbhofgesetz Sept. 1933)",
     "Conservative nationalist proposal to protect family farms from division "
     "through inheritance and from forced sale through debt. This bill served "
     "as a conceptual precursor to the Nazi Reichserbhofgesetz (Hereditary "
     "Farm Law) of September 1933.",
     "Protection of farms from subdivision through inheritance; restrictions "
     "on farm sale to non-farmers; hereditary farm designation.",
     "Family farmers, peasant landholders",
     "DNVP (Deutschnationale Volkspartei) / Stahlhelm"),
]


# ── Create workbook ──────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Weimar Ag Bills 1919-1933"

# Column headers
headers = [
    "Year",
    "German Name (Deutscher Name)",
    "English Name",
    "Date Enacted / Proposed",
    "RGBl. Citation",
    "Category",
    "Status (Passed / Rejected)",
    "Description",
    "Key Provisions",
    "Primary Beneficiaries",
    "Initiating Body",
]

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

passed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
rejected_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
decree_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

cell_alignment = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
for row_num, bill in enumerate(bills, 2):
    (year, german_name, english_name, date_enacted, rgbl, category,
     status, description, provisions, beneficiaries, initiator) = bill

    row_data = [year, german_name, english_name, date_enacted, rgbl,
                category, status, description, provisions, beneficiaries, initiator]

    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Colour-code by status
        status_lower = status.lower()
        if "rejected" in status_lower or "not passed" in status_lower or "failed" in status_lower:
            cell.fill = rejected_fill
        elif "emergency" in status_lower or "decree" in status_lower:
            cell.fill = decree_fill
        elif "passed" in status_lower:
            cell.fill = passed_fill

# Set column widths
col_widths = [8, 55, 50, 22, 25, 35, 30, 70, 70, 45, 45]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Row height for header
ws.row_dimensions[1].height = 35

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(bills) + 1}"

# ── Summary sheet ────────────────────────────────────────────────────────────

ws2 = wb.create_sheet("Summary & Legend")

ws2["A1"] = "Weimar Republic (1919-1933) Agricultural Bills Summary"
ws2["A1"].font = Font(bold=True, size=14)

ws2["A3"] = "This dataset covers parliamentary bills, laws, ordinances, and emergency decrees"
ws2["A4"] = "related to agricultural protectionism, public rural credits, mortgage credits,"
ws2["A5"] = "land reform, and related beneficial agricultural legislation passed or proposed"
ws2["A6"] = "during the Weimar Republic period (1919-1933)."

ws2["A8"] = "Colour Legend:"
ws2["A8"].font = Font(bold=True, size=11)

ws2["A9"] = "Green"
ws2["A9"].fill = passed_fill
ws2["B9"] = "Passed (regular legislation)"

ws2["A10"] = "Orange"
ws2["A10"].fill = decree_fill
ws2["B10"] = "Passed (emergency decree / Notverordnung under Article 48)"

ws2["A11"] = "Red/Pink"
ws2["A11"].fill = rejected_fill
ws2["B11"] = "Rejected / Not passed / Failed"

ws2["A13"] = "Categories:"
ws2["A13"].font = Font(bold=True, size=11)

categories = [
    "Land Reform / Settlement – Laws governing land redistribution and agricultural settlement",
    "Rural Credit – Laws establishing or regulating agricultural credit institutions",
    "Agricultural Protectionism / Tariffs – Import duties and trade barriers for agricultural goods",
    "Agricultural Debt / Mortgage Relief – Debt moratoriums, revaluation, and foreclosure protection",
    "Agricultural Finance / Emergency Measures – Broad emergency decrees with agricultural provisions",
    "Agricultural Market Regulation – Price controls, market intervention, and trade regulation",
    "Agricultural Labour – Regulations governing farm workers' conditions and rights",
    "Regional Aid – Geographically targeted aid programmes (e.g., Osthilfe)",
]

for i, cat in enumerate(categories):
    ws2[f"A{14 + i}"] = f"• {cat}"

ws2[f"A{14 + len(categories) + 1}"] = "Total bills documented:"
ws2[f"A{14 + len(categories) + 1}"].font = Font(bold=True)
ws2[f"B{14 + len(categories) + 1}"] = len(bills)

passed_count = sum(1 for b in bills if "passed" in b[6].lower() and "rejected" not in b[6].lower())
rejected_count = sum(1 for b in bills if "rejected" in b[6].lower() or "not passed" in b[6].lower() or "failed" in b[6].lower())

ws2[f"A{14 + len(categories) + 2}"] = "Passed (including emergency decrees):"
ws2[f"A{14 + len(categories) + 2}"].font = Font(bold=True)
ws2[f"B{14 + len(categories) + 2}"] = passed_count

ws2[f"A{14 + len(categories) + 3}"] = "Rejected / Not passed:"
ws2[f"A{14 + len(categories) + 3}"].font = Font(bold=True)
ws2[f"B{14 + len(categories) + 3}"] = rejected_count

ws2[f"A{14 + len(categories) + 5}"] = "Key Sources:"
ws2[f"A{14 + len(categories) + 5}"].font = Font(bold=True, size=11)

sources = [
    "Reichsgesetzblatt (RGBl) 1919-1933 – Official law gazette of the German Reich",
    "JSTOR: Agrarian Protectionism in the Weimar Republic (https://www.jstor.org/stable/260171)",
    "Oxford Handbook of the Weimar Republic – Agriculture and Rural Society chapter",
    "Foreign Affairs (1932): The Crisis in German Agriculture",
    "Gesetze-im-Internet.de: RSiedlG (Reichssiedlungsgesetz), PachtkredG (Pachtkreditgesetz)",
    "Historisches Lexikon Bayerns: Osthilfe, 1926-1937",
    "1000dokumente.de: Verordnung über die Errichtung der Deutschen Rentenbank",
    "Wikipedia: Reich Settlement Law, Reich Ministry of Food and Agriculture",
    "Gerschenkron, A.: Bread and Democracy in Germany (1943)",
    "Moeller, R.: Peasants and Lords in Modern Germany (1986)",
    "Peukert, D.: The Weimar Republic (1991)",
]

for i, src in enumerate(sources):
    ws2[f"A{14 + len(categories) + 6 + i}"] = f"• {src}"

ws2.column_dimensions["A"].width = 80
ws2.column_dimensions["B"].width = 50

# Save
output_path = "/home/runner/work/landpolicies/landpolicies/weimar_agricultural_bills_1919_1933.xlsx"
wb.save(output_path)
print(f"Excel file created: {output_path}")
print(f"Total bills: {len(bills)}")
print(f"  Passed: {passed_count}")
print(f"  Rejected/Not passed: {rejected_count}")
